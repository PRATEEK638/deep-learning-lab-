"""excel_tool – safe Excel / CSV operations via openpyxl for FRIDAY.

Supported actions:
  read_sheet   – read a worksheet into list-of-rows (respects max_rows)
  write_sheet  – write list-of-rows to a worksheet (copy-on-write safe)
  list_sheets  – list worksheet names
  append_rows  – append rows to an existing sheet
  summary      – return basic stats (shape, column names, first 5 rows)
"""

from __future__ import annotations

import copy
import pathlib
import shutil
import tempfile
from typing import Any

from ..config import get_config
from ..policy import PolicyEngine
from ..tools import BaseTool, SafetyLevel, SideEffect, ToolSchema


class ExcelTool(BaseTool):
    """Excel/XLSX operations via openpyxl with row-count and path guardrails."""

    _SCHEMA = ToolSchema(
        name="excel_tool",
        description=(
            "Read or write Excel (.xlsx) files safely.  "
            "Supports read_sheet, write_sheet, list_sheets, append_rows, summary."
        ),
        side_effect=SideEffect.WRITE,
        safety_level=SafetyLevel.LOCAL_ONLY,
        timeout_seconds=60,
        retries=1,
        requires_confirmation=False,
        dry_run=False,
        idempotent=False,
        input_schema={
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": [
                        "read_sheet", "write_sheet", "list_sheets",
                        "append_rows", "summary",
                    ],
                },
                "path": {"type": "string"},
                "sheet": {"type": "string"},
                "rows": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "Rows for write_sheet / append_rows",
                },
                "header": {
                    "type": "boolean",
                    "default": True,
                    "description": "Whether the first row is a header",
                },
            },
            "required": ["action", "path"],
        },
        output_schema={
            "type": "object",
            "properties": {
                "result": {},
                "message": {"type": "string"},
            },
        },
    )

    def __init__(self, policy: PolicyEngine | None = None, config: dict | None = None):
        try:
            import openpyxl  # noqa: F401
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for excel_tool. "
                "Install it with: pip install openpyxl"
            ) from exc

        cfg = config or get_config()
        self._policy = policy or PolicyEngine(cfg)
        self._max_rows: int = cfg.get("policy", {}).get("max_excel_rows", 100_000)

    @property
    def schema(self) -> ToolSchema:
        return self._SCHEMA

    def _check_path(self, path: str) -> pathlib.Path:
        decision = self._policy.check_path(path)
        if not decision.allowed:
            raise PermissionError(decision.reason)
        p = pathlib.Path(path).expanduser()
        if p.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            raise ValueError(f"Unsupported file type: {p.suffix}")
        return p

    def _load_wb(self, path: pathlib.Path):
        import openpyxl
        return openpyxl.load_workbook(path)

    def _list_sheets(self, path: str) -> list[str]:
        p = self._check_path(path)
        wb = self._load_wb(p)
        return wb.sheetnames

    def _read_sheet(self, path: str, sheet: str | None, header: bool) -> dict:
        import openpyxl
        p = self._check_path(path)
        wb = self._load_wb(p)
        ws = wb[sheet] if sheet else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > self._max_rows:
            raise ValueError(
                f"Sheet has {len(rows)} rows, exceeding max {self._max_rows}"
            )
        if header and rows:
            return {"columns": list(rows[0]), "rows": [list(r) for r in rows[1:]]}
        return {"columns": None, "rows": [list(r) for r in rows]}

    def _write_sheet(self, path: str, sheet: str | None, rows: list[list]) -> str:
        import openpyxl
        p = self._check_path(path)
        if len(rows) > self._max_rows:
            raise ValueError(
                f"Attempting to write {len(rows)} rows, exceeding max {self._max_rows}"
            )
        # Copy-on-write safety: write to tmp then atomically replace
        if p.exists():
            wb = self._load_wb(p)
        else:
            wb = openpyxl.Workbook()

        ws_name = sheet or (wb.active.title if wb.active else "Sheet1")
        if ws_name in wb.sheetnames:
            ws = wb[ws_name]
            # Clear existing content
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(ws_name)

        for row in rows:
            ws.append(row)

        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", dir=p.parent, delete=False
        ) as tmp:
            tmp_path = pathlib.Path(tmp.name)

        try:
            wb.save(tmp_path)
            shutil.move(str(tmp_path), str(p))
        except Exception:
            tmp_path.unlink(missing_ok=True)
            raise

        return str(p)

    def _append_rows(self, path: str, sheet: str | None, rows: list[list]) -> str:
        import openpyxl
        p = self._check_path(path)
        if not p.exists():
            raise FileNotFoundError(f"'{p}' does not exist; use write_sheet to create it")

        wb = self._load_wb(p)
        ws = wb[sheet] if sheet else wb.active
        current_rows = ws.max_row or 0
        if current_rows + len(rows) > self._max_rows:
            raise ValueError(
                f"Appending {len(rows)} rows would exceed max {self._max_rows}"
            )
        for row in rows:
            ws.append(row)

        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", dir=p.parent, delete=False
        ) as tmp:
            tmp_path = pathlib.Path(tmp.name)

        try:
            wb.save(tmp_path)
            shutil.move(str(tmp_path), str(p))
        except Exception:
            tmp_path.unlink(missing_ok=True)
            raise

        return str(p)

    def _summary(self, path: str, sheet: str | None) -> dict:
        data = self._read_sheet(path, sheet, header=True)
        preview = data["rows"][:5]
        return {
            "columns": data["columns"],
            "total_rows": len(data["rows"]),
            "preview": preview,
        }

    def _execute(
        self,
        action: str,
        path: str,
        sheet: str | None = None,
        rows: list | None = None,
        header: bool = True,
        **_,
    ) -> Any:
        if action == "list_sheets":
            return self._list_sheets(path)
        if action == "read_sheet":
            return self._read_sheet(path, sheet, header)
        if action == "write_sheet":
            if rows is None:
                raise ValueError("'rows' is required for write_sheet")
            return self._write_sheet(path, sheet, rows)
        if action == "append_rows":
            if rows is None:
                raise ValueError("'rows' is required for append_rows")
            return self._append_rows(path, sheet, rows)
        if action == "summary":
            return self._summary(path, sheet)
        raise ValueError(f"Unknown action: {action!r}")
